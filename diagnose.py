import sys
import os
import openpyxl
import psycopg2

DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("jdbc:"):
    DATABASE_URL = DATABASE_URL.removeprefix("jdbc:")

XLSX_PATH = sys.argv[1]

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb["Вопросы"]

header_row = [cell.value for cell in ws[1]]
col_index = {name: idx for idx, name in enumerate(header_row) if name}

excel_keys_by_test = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    key = row[col_index["question_key"]]
    test_number = row[col_index["test_number"]]
    if key:
        excel_keys_by_test.setdefault(test_number, []).append(str(key).strip())

conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()
cur.execute("SELECT question_key, test_number FROM questions ORDER BY test_number, id")
db_rows = cur.fetchall()
cur.close()
conn.close()

db_keys_by_test = {}
for key, test_number in db_rows:
    db_keys_by_test.setdefault(test_number, []).append(key)

for test_number in sorted(set(list(excel_keys_by_test.keys()) + list(db_keys_by_test.keys()))):
    excel_keys = excel_keys_by_test.get(test_number, [])
    db_keys = db_keys_by_test.get(test_number, [])
    print(f"\n=== Тест {test_number} ===")
    print(f"В Excel строк: {len(excel_keys)}, в базе строк: {len(db_keys)}")

    only_in_db = set(db_keys) - set(excel_keys)
    only_in_excel = set(excel_keys) - set(db_keys)

    if only_in_db:
        print(f"Есть в базе, но НЕТ в Excel (осиротевшие записи): {only_in_db}")
    if only_in_excel:
        print(f"Есть в Excel, но нет в базе (ещё не импортированы): {only_in_excel}")