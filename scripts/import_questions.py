import sys
import os
import openpyxl
import psycopg2

DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("jdbc:"):
    DATABASE_URL = DATABASE_URL.removeprefix("jdbc:")
if not DATABASE_URL:
    print("Переменная DATABASE_URL не задана")
    sys.exit(1)

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "Вопросы_ПДД_5_языков.xlsx"

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb["Вопросы"]

header_row = [cell.value for cell in ws[1]]
col_index = {name: idx for idx, name in enumerate(header_row) if name}

required = ["test_number", "image_filename", "correct_option", "question_key"]
for col_name in required:
    if col_name not in col_index:
        print(f"В таблице нет колонки {col_name}")
        sys.exit(1)

langs = ["ru", "en", "es", "de", "uk"]

conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()

inserted = 0
updated = 0
skipped = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    def get(col_name):
        idx = col_index.get(col_name)
        return row[idx] if idx is not None else None

    test_number = get("test_number")
    image_filename = get("image_filename")
    correct_option = get("correct_option")
    question_key = get("question_key")
    question_text_ru = get("question_text_ru")

    if not test_number or not image_filename or not question_text_ru or not question_key:
        skipped += 1
        continue

    filename = str(image_filename).strip()
    if not filename.lower().endswith(".jpg"):
        filename += ".jpg"

    key = str(question_key).strip()

    values_by_lang = {}
    for lang in langs:
        values_by_lang[lang] = {
            "question_text": get(f"question_text_{lang}"),
            "option_a": get(f"option_a_{lang}"),
            "option_b": get(f"option_b_{lang}"),
            "option_c": get(f"option_c_{lang}"),
            "option_d": get(f"option_d_{lang}"),
            "explanation": get(f"explanation_{lang}"),
        }

    cur.execute("SELECT id FROM questions WHERE question_key = %s", (key,))
    existing = cur.fetchone()

    columns = ["test_number", "image_filename", "correct_option",
               "question_text", "option_a", "option_b", "option_c", "option_d", "explanation"]
    values = [test_number, filename, correct_option,
              values_by_lang["ru"]["question_text"], values_by_lang["ru"]["option_a"],
              values_by_lang["ru"]["option_b"], values_by_lang["ru"]["option_c"],
              values_by_lang["ru"]["option_d"], values_by_lang["ru"]["explanation"]]

    for lang in ["en", "es", "de", "uk"]:
        columns += [f"question_text_{lang}", f"option_a_{lang}", f"option_b_{lang}",
                    f"option_c_{lang}", f"option_d_{lang}", f"explanation_{lang}"]
        values += [values_by_lang[lang]["question_text"], values_by_lang[lang]["option_a"],
                   values_by_lang[lang]["option_b"], values_by_lang[lang]["option_c"],
                   values_by_lang[lang]["option_d"], values_by_lang[lang]["explanation"]]

    if existing:
        set_clause = ", ".join(f"{col} = %s" for col in columns)
        cur.execute(f"UPDATE questions SET {set_clause} WHERE question_key = %s", values + [key])
        updated += 1
    else:
        cur.execute(
            f"INSERT INTO questions (question_key, {', '.join(columns)}) VALUES (%s, {', '.join(['%s'] * len(columns))})",
            [key] + values
        )
        inserted += 1

conn.commit()
cur.close()
conn.close()

print(f"Добавлено: {inserted}, обновлено: {updated}, пропущено: {skipped}")