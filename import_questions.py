import sys
import os
import openpyxl
import psycopg2

DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("jdbc:"):
    DATABASE_URL = DATABASE_URL.removeprefix("jdbc:")
if not DATABASE_URL:
    print("Переменная DATABASE_URL не задана")
    sys.exit(1)

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "Вопросы_ПДД_5_языков.xlsx"

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb["Вопросы"]

header_row = [cell.value for cell in ws[1]]
col_index = {name: idx for idx, name in enumerate(header_row) if name}

required = ["test_number", "image_filename", "question_text_ru", "option_a_ru",
            "option_b_ru", "option_c_ru", "option_d_ru", "correct_option", "explanation_ru", "question_key"]
missing = [c for c in required if c not in col_index]
if missing:
    print(f"В таблице нет колонок: {missing}")
    print(f"Реально найденные заголовки: {list(col_index.keys())}")
    sys.exit(1)

conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()

inserted = 0
updated = 0
skipped = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    def get(col_name):
        return row[col_index[col_name]]

    test_number = get("test_number")
    image_filename = get("image_filename")
    question_text = get("question_text_ru")
    option_a = get("option_a_ru")
    option_b = get("option_b_ru")
    option_c = get("option_c_ru")
    option_d = get("option_d_ru")
    correct_option = get("correct_option")
    explanation = get("explanation_ru")
    question_key = get("question_key")

    if not test_number or not image_filename or not question_text or not question_key:
        skipped += 1
        continue

    filename = str(image_filename).strip()
    if not filename.lower().endswith(".jpg"):
        filename += ".jpg"

    key = str(question_key).strip()

    cur.execute("SELECT id FROM questions WHERE question_key = %s", (key,))
    existing = cur.fetchone()

    if existing:
        cur.execute("""
            UPDATE questions
            SET test_number = %s, image_filename = %s, question_text = %s, option_a = %s, option_b = %s,
                option_c = %s, option_d = %s, correct_option = %s, explanation = %s
            WHERE question_key = %s
        """, (test_number, filename, question_text, option_a, option_b, option_c, option_d, correct_option, explanation, key))
        updated += 1
    else:
        cur.execute("""
            INSERT INTO questions
            (question_key, test_number, image_filename, question_text, option_a, option_b, option_c, option_d, correct_option, explanation)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (key, test_number, filename, question_text, option_a, option_b, option_c, option_d, correct_option, explanation))
        inserted += 1

conn.commit()
cur.close()
conn.close()

print(f"Добавлено: {inserted}, обновлено: {updated}, пропущено: {skipped}")