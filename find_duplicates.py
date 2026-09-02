import sys
import openpyxl
from collections import defaultdict

XLSX_PATH = sys.argv[1]

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb["Вопросы"]

header_row = [cell.value for cell in ws[1]]
filename_col = header_row.index("image_filename")

seen = defaultdict(list)
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    filename = row[filename_col]
    if filename:
        seen[str(filename).strip()].append(row_idx)

for filename, rows in seen.items():
    if len(rows) > 1:
        print(f"'{filename}' встречается в строках: {rows}")